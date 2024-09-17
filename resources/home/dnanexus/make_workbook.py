import dxpy
import networkx
import gzip
import json
import obonet
import openpyxl.drawing
import openpyxl.drawing.image
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, DEFAULT_FONT, Font
import os
from os import path
from pathlib import Path
import re
import dxpy

from excel_styles import ExcelStyles, DropDown
from get_variant_info import VariantUtils, VariantNomenclature

DEFAULT_FONT.name = 'Calibri'
# row and col counts that are to be unlocked next to
# populated table in all sheets if it is dias pipeline
# required for 'lock_sheet' function
ROW_TO_UNLOCK = 500
COL_TO_UNLOCK = 200


class excel():
    '''
    Functions to generate excel workbook of variants
    '''
    def __init__(self, args) -> None:
        self.args = args
        self.wgs_data = None
        self.writer = None
        self.workbook = None
        self.proband = None
        self.proband_sex = None
        self.mother = None
        self.father = None
        self.other_relation = False
        self.bold_content = None
        self.summary_content = None
        self.mane = None
        self.refseq_tsv = None
        self.config = None
        self.gel_index = None
        self.ex_index = None
        self.var_df = None
        self.genome_format = None
        self.genome_data_format = None
        self.column_list = [
            "Gene",
            "Chr",
            "Pos",
            "End",
            "Length",
            "Ref",
            "Alt",
            'STR1',
            'STR2',
            "Repeat",
            "Copy Number",
            "Type",
            "HGVSc",
            "HGVSp",
            "Priority",
            "Zygosity",
            "Panel MOI",
            "Inheritance",
            "Depth",
            "AF Max",
            "Penetrance filter",
            "Comment",
            "Checker comment"
        ]

    def generate(self) -> None:
        """
        Calls all methods in excel() to generate output file.
        """
        # Initiate files and workbook to write in
        self.open_files()
        self.writer = pd.ExcelWriter(
            self.args.output_filename, engine='openpyxl'
        )
        self.workbook = self.writer.book
        print(f"Writing to {self.args.output_filename}...")
        # Write in workbook
        self.summary_page()
        self.get_interpreted_genome_format()
        self.index_interpretation_services()
        self.create_gel_tiering_variant_page()
        self.create_additional_analysis_page()
        self.str_image_page()
        self.writer.close()
        if self.args.acmg:
            for i in range(1, self.args.acmg+1):
                self.write_snv_reporting_template(i)
        if self.args.cnv:
            for i in range(1, self.args.cnv+1):
                self.write_cnv_reporting_template(i)
        self.workbook.save(self.args.output_filename)
        if self.args.acmg:
            DropDown.drop_down(self)
        print('Done!')

    def open_files(self):
        '''
        Open input files and read into variables.
        '''
        with open(self.args.json) as f:
            self.wgs_data = json.load(f)

        with gzip.open(self.args.mane_file) as f:
            self.mane = [x.decode('utf8').strip() for x in f.readlines()]

        with gzip.open(self.args.refseq_tsv) as f:
            self.refseq_tsv = [x.decode('utf8').strip() for x in f.readlines()]

        with open(self.args.config) as fh:
            self.config = json.load(fh)
        
        with open(self.args.panels) as f:
            self.panels = json.load(f)

    def summary_page(self):
        '''
        Add summary page. Create a page in the workbook to populate with
        details about the case and variants for interpretation.
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        '''
        summary_sheet = self.workbook.create_sheet("Summary")

        self.bold_content = {
            (1, 1): "Family ID",
            (6, 1): "Proband",
            (7, 1): "Mother",
            (8, 1): "Father",
            (5, 2): "ID",
            (5, 3): "GM/SP number",
            (5, 4): "NUH number",
            (5, 5): "Sex",
            (5, 6): "Affected?",
            (5, 7): "HPO",
            (2, 1): "Clinical indication",
            (1, 8): "Flags",
            (3, 1): "Penetrance",
            (30, 1): "Overall result",
            (31, 1): "Confirmation",
            (33, 2): "Name",
            (33, 3): "Date",
            (34, 1): "Primary analysis",
            (35, 1): "Data check",
            (10, 1): "LP number",
            (12, 1): "Panels",
            (13, 1): "Panel ID",
            (13, 2): "Test code",
            (13, 3): "Panel name and version",
            (21, 2): "In this case",
            (21, 3): "To be reported",
            (22, 1): "SNV Tier 1",
            (23, 1): "SNV Tier 2",
            (24, 1): "CNV Tier 1",
            (25, 1): "STR Tier 1",
            (27, 1): "Exomiser top 3 (score ≥ 0.75)",
            (28, 1): "De novo",
        }

        self.summary_content = {
            (1, 9): str(
                self.wgs_data[
                    "interpretation_request_data"
                ]['json_request']['interpretationFlags']
            ),
            (1, 2): self.wgs_data["family_id"],
        }

        # Add panel data, penetrance data and data about family members
        self.get_panels()
        self.get_penetrance()
        self.person_data()
        if self.args.epic_clarity:
            self.add_epic_data()

        # write summary content and titles to page
        for key, val in self.bold_content.items():
            summary_sheet.cell(key[0], key[1]).value = val
            summary_sheet.cell(key[0], key[1]).font = Font(
                bold=True, name=DEFAULT_FONT.name
            )
        for key, val in self.summary_content.items():
            summary_sheet.cell(key[0], key[1]).value = val

        # Set column widths
        summary_sheet.column_dimensions["A"].width = 20
        summary_sheet.column_dimensions["B"].width = 16
        for col in ["C", "D", "F"]:
            summary_sheet.column_dimensions[col].width = 14

        row_ranges = {
            'horizontal': [
                'A33:C33', 'A34:C34', 'A36:C36', 'A30:B30', 'A31:B31',
                'A32:B32'
            ],
            'vertical': [
                'A33:A35', 'B33:B35', 'C33:C35', 'D33:D35', 'A30:A31',
                'B30:B31', 'C30:C31',
            ]
        }

        ExcelStyles.borders(self, row_ranges, summary_sheet)

    def get_hpo_obo(self):
        '''
        Select which version of HPO from the input config to use based on
        which version was used by GEL when the JSON was made.
        Inputs:
            version (str): version of HPO used in JSON
        Outputs:
            None, downloads the HPO obo for the version of HPO used in the GEL
            JSON
        '''
        # Get HPO version from JSON
        obo = None
        version = self.wgs_data[
                    "interpretation_request_data"
                ]['json_request']["pedigree"][
            "members"
            ][0]["hpoTermList"][0]['hpoBuildNumber']

        # Find dx file ID in config for obo file for that version of HPO
        for k, v in self.config['obos'].items():
            if k == version:
                obo = v

        # If no match found error, else download the HPO obo + call it hpo.obo
        if obo is None:
            raise RuntimeError(
                f"HPO version in JSON {version} not found in config\n"
                f"{self.config}"
            )

        dxpy.download_dxfile(obo, "hpo.obo")

    def get_hpo_terms(self, member):
        '''
        Use obo hpo term ontology (.obo) file to convert HPO IDs to names.
        This file has been downloaded and named hpo.obo by get_hpo_obo
        Inputs:
            member (dict): family member dictionary extracted from GEL JSON
        Outputs:
            hpo_names (list): list of HPO names corresponding to HPO terms for
            that family member
        '''

        hpo_terms = []
        hpo_names = []
        if member["hpoTermList"]:
            graph = obonet.read_obo("hpo.obo")
            # Read in HPO IDs from JSON
            for i in member["hpoTermList"]:
                hpo_terms.append(i["term"])
            # Convert term to name using obo file
            for term in hpo_terms:
                hpo_dict = graph.nodes[term]
                hpo_name = hpo_dict['name']
                hpo_names.append(hpo_name)

            hpo_names = '; '.join(hpo_names)

        else:
            hpo_names = None

        return hpo_names

    def add_person_data_to_summary(self, member, index):
        '''
        Function to add data that is added for all family members to the
        summary sheet. Isolated here as it is the same for all family members
        Inputs:
            member (dict): dictionary for that member in the pedigree in the
            GEL JSON
            index (int): row to populate in the sheet, specific to that family
            member
        Outputs:
            None, adds content to openpxyl workbook
        '''
        self.summary_content[(index, 2)] = member["participantId"]
        self.summary_content[(index, 5)] = member["sex"]
        self.summary_content[(index, 6)] = member["affectionStatus"]
        self.summary_content[(index, 7)] = self.get_hpo_terms(member)

    def person_data(self):
        '''
        Find data for participants and add to summary sheet.
        This function will find the proband and add their affected status, HPO
        term names, participant ID, sample ID to the summary sheet
        It will also do this for any relatives in the JSON
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        '''
        # Get version of HPO to use for terms
        self.get_hpo_obo()
        pb_relate = lambda x: x["additionalInformation"]["relation_to_proband"]

        for member in self.wgs_data[
                    "interpretation_request_data"
                ]['json_request']["pedigree"]["members"]:

            # Ignore participants starting NR, as they have no sequence data
            if member["participantId"].startswith('NR'):
                continue

            if member["isProband"]:
                self.add_person_data_to_summary(member, 6)
                self.proband = member["participantId"]
                self.proband_sex = member["sex"]
                self.summary_content[(10, 2)] = member["samples"][0][
                    "sampleId"
                ]

            elif pb_relate(member) == "Mother":
                self.add_person_data_to_summary(member, 7)
                self.mother = member["participantId"]

            elif pb_relate(member) == "Father":
                self.add_person_data_to_summary(member, 8)
                self.father = member["participantId"]

            else:
                self.summary_content[(9, 1)] = pb_relate(member)
                self.add_person_data_to_summary(member, 9)
                self.other_relation = True

    def get_panels(self):
        '''
        Function to get panels from JSON and add to summary content dict to
        then add to summary content sheet
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        '''
        indications = []
        row = 14
        for panel in self.wgs_data[
                "interpretation_request_data"
            ]['json_request']['pedigree']['analysisPanels']:
            indications.append(panel['specificDisease'])
            panel_details = self.panels.get(panel['panelId'])
            panel_version = panel['panelVersion']
            self.summary_content[(row, 1)] = panel['panelId']
            self.summary_content[(row, 2)] = panel_details.get('rcode')
            self.summary_content[(row, 3)] = panel_details.get(
                'panel_name'
            ) + f' ({panel_version})'

            row += 1

       # Add clinical indication content
        self.summary_content[(2, 2)] = ", ".join(set(indications))


    def get_penetrance(self):
        '''
        Get penetrance info from JSON and add to summary content dict to
        then add to summary content sheet
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        '''
        p_list = []
        for penetrance in self.wgs_data[
                    "interpretation_request_data"
                ]['json_request']['pedigree']['diseasePenetrances']:
            p_list.append(penetrance['penetrance'])

        disease_penetrance = ', '.join(p_list)

        self.summary_content[(3, 2)] = disease_penetrance

    def add_epic_data(self):
        '''
        Read in data from Epic Clarity extract and add to summary page,
        This function assumes that in a case where there is a proband and
        parent(s), the youngest person is the proband, the older female is the
        mother and the older male is the father.
        It does not work for cases where there are siblings (unclear which is
        the proband), or other relations
        Inputs:
            None, uses Epic clarity export
        Outputs:
            None, adds Epic data to Excel workbook.
        '''
        family_id = self.wgs_data['family_id']
        # Only run if there are only parents and proband
        if self.other_relation is False:
            # Read in xlsx as df, using only relevant columns
            df = pd.read_excel(
                self.args.epic_clarity,
                usecols=[
                    "WGS Referral ID",
                    "External Specimen Identifier",
                    "Specimen Identifier",
                    "Sex",
                    "YOB"
                    ]
            )
            # Filter df to only have rows with the family ID for this case
            fam_df = df.loc[df['WGS Referral ID'] == family_id]

            if not fam_df.empty:
            # Use most recent year of birth to work out proband, then get IDs
                pb_idx = fam_df['YOB'].idxmax()
                pb_age = fam_df['YOB'].max()
                pb_sp, pb_nuh = self.get_ids(fam_df, pb_idx)

                m_sp, m_nuh = self.get_parent_ids(pb_age, fam_df, "FEMALE")
                f_sp, f_nuh = self.get_parent_ids(pb_age, fam_df, "MALE")

                self.summary_content[(6, 3)] = pb_sp
                self.summary_content[(6, 4)] = pb_nuh

                self.summary_content[(7, 3)] = m_sp
                self.summary_content[(7, 4)] = m_nuh

                self.summary_content[(8, 3)] = f_sp
                self.summary_content[(8, 4)] = f_nuh
            else:
                print(
                    f"Family ID {family_id} not found in Epic Clarity export. "
                    "Continuing without adding sample IDs from Epic..."
                )
        else:
            print(
                "Cannot reliably determine family relationships based on age"
                " and sex due to the presence of family members who are not"
                " the proband or parent(s). Continuing without adding sample "
                "IDs from Epic..."
            )

    @staticmethod
    def get_ids(df, row):
        '''
        Get the SP number and NUH ID from a given row in the dataframe
        Inputs
            df: pandas dataframe of the family in the Epic clarity export
            row: row in the df for which to get data
        Outputs
            sp_number: Epic sample number
            nuh_id: External sample ID for NUH samples
        '''
        sp_number = df.loc[row, "Specimen Identifier"]
        nuh_id = df.loc[row, "External Specimen Identifier"]
        return sp_number, nuh_id

    def get_parent_ids(self, pb_yob, df, sex):
        '''
        Get the SP number and NUH ID for the parents, assuming that mother is
        female and older than the proband, and father is male and older than
        the proband
        Inputs
            df: pandas dataframe of the family in the Epic clarity export
            pb_yob: year of birth of the proband
            sex: sex of the parent
        Outputs:
            p_sp: Epic sample number for the parent
            p_nuh: External sample ID for NUH samples for the parent
        '''
        parent_df = df[(df['Sex'] == sex) & (df['YOB'] < pb_yob)]
        if parent_df.empty:
            p_sp = None
            p_nuh = None
        else:
            p_idx = parent_df.index.tolist()[0]
            p_sp, p_nuh = self.get_ids(df, p_idx)
        return p_sp, p_nuh

    def get_interpreted_genome_format(self):
        '''
        There are two versions of the formatting for interpreted genomes fields
        one is interpretedGenomes > interpretedGenomeData and the other is
        interpreted_genome > interpreted_genome_data. This function finds out
        which to use.
        Inputs:
            None
        Outputs:
            None, sets self.genome_format and self.genome_data_format for use
            in creating variant pages
        '''
        if self.wgs_data.get('interpretedGenomes') is not None:
            self.genome_format = 'interpretedGenomes'
            self.genome_data_format = 'interpretedGenomeData'
        elif self.wgs_data.get('interpreted_genome') is not None:
            self.genome_format = 'interpreted_genome'
            self.genome_data_format = 'interpreted_genome_data'
        else:
            raise RuntimeError(
                "JSON does not have interpreted_genome or interpretedGenomes "
                "field which is needed to pull out variants."
            )

    def index_interpretation_services(self):
        '''
        The JSON contains two interpretation services: GEL tiering and Exomiser
        This function finds the index for each so these can be referred to
        correctly and sets self.gel_index to the index for the GEL tiering and
        self.ex_index to to the index for the Exomiser interpretation.
        Inputs:
            None
        Outputs:
            None, sets indexs for GEL tiering and exomiser in the JSON.
        '''
        for interpretation in self.wgs_data[self.genome_format]:
            if interpretation[self.genome_data_format][
                'interpretationService'
                ] == 'genomics_england_tiering':
                self.gel_index = self.wgs_data[
                    self.genome_format
                    ].index(interpretation)
            elif interpretation[self.genome_data_format][
                'interpretationService'
                ] == 'Exomiser':
                self.ex_index = self.wgs_data[
                    self.genome_format
                    ].index(interpretation)
            else:
                raise RuntimeError(
                    "Interpretation services in JSON not recognised as "
                    "'genomics_england_tiering' or 'Exomiser'"
                )

    def str_image_page(self):
        '''
        STR table is useful for interpretation, so will be included on a sheet
        so it can be referred to during interpretation.
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        '''
        str_sheet = self.workbook.create_sheet("STR guidelines")
        script_dir = os.path.dirname(__file__)
        img_folder = "images/str_table.png"
        img_path = os.path.join(script_dir, img_folder)
        img = openpyxl.drawing.image.Image(img_path)
        img.anchor = 'B4'
        str_sheet.add_image(img)
        str_sheet['B2'] = (
            "From CU-WG-REF-40 Guidelines for Rare Disease Whole Genome "
            "Sequencing & Next Generation Sequencing Panel Interpretation & "
            "Reporting"
        )

    def create_gel_tiering_variant_page(self):
        '''
        Take variants from GEL tiering JSON and format into sheet in Excel
        workbook.
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        '''
        variant_list = []

        # SNVs
        for snv in self.wgs_data[self.genome_format][
            self.gel_index
            ][self.genome_data_format]["variants"]:
            for event in snv["reportEvents"]:
                if event["tier"] in ["TIER1", "TIER2"]:
                    event_index = snv["reportEvents"].index(event)
                    var_dict = VariantUtils.get_snv_info(
                        snv,
                        self.proband,
                        event_index,
                        self.column_list,
                        self.mother,
                        self.father,
                        self.proband_sex
                        )
                    c_dot, p_dot = VariantNomenclature.get_hgvs_gel(
                        snv,
                        self.mane,
                        self.refseq_tsv
                        )
                    var_dict["HGVSc"] = c_dot
                    var_dict["HGVSp"] = p_dot
                    variant_list.append(var_dict)

        # STRs
        for s_t_r in self.wgs_data[self.genome_format][
            self.gel_index
            ][self.genome_data_format][
                "shortTandemRepeats"
            ]:
            for event in s_t_r["reportEvents"]:
                if event["tier"] == "TIER1":
                    var_dict = VariantUtils.get_str_info(
                        s_t_r, self.proband, self.column_list
                    )
                    variant_list.append(var_dict)

        # CNVs
        for cnv in self.wgs_data[self.genome_format][
            self.gel_index
            ][self.genome_data_format]["structuralVariants"]:
            for event in cnv["reportEvents"]:
                event_index = cnv["reportEvents"].index(event)
                # CNVs can be reported as Tier 1 or Tier A, GEL updated the
                # nomenclature in 2024
                if cnv["reportEvents"][event_index]["tier"] in [
                    "TIER1", "TIERA"
                    ]:
                    var_dict = VariantUtils.get_cnv_info(
                        cnv, event_index, self.column_list
                    )
                    variant_list.append(var_dict)

        # Add all variants into dataframe
        self.var_df = pd.DataFrame(variant_list)
        self.var_df = self.var_df.drop_duplicates()

        # Prepare to add counts to summary sheet
        summary_sheet = self.workbook["Summary"]
        count_dict = {
            'B22': "TIER1_SNV",
            'B23': "TIER2_SNV",
            'B24': "TIER1_CNV",
            'B25': "TIER1_STR",
        }

        # if df is not empty, sort and add counts of each variant type to
        # summary sheet
        if not self.var_df.empty:
            self.var_df['Depth'] = self.var_df['Depth'].astype(object)

            # if variant is in both Tier 1 and Tier 2, keep Tier 1 entry only
            self.var_df = self.var_df.sort_values(
                by=['Priority']
                ).drop_duplicates(subset=['Chr', 'Pos', 'Ref', 'Alt', 'End'])

            # Sort by Priority and then Gene symbol
            self.var_df = self.var_df.sort_values(['Priority', 'Gene'])

            # Add variant counts to summary sheet
            for key, val in count_dict.items():
                if val in self.var_df.Priority.values:
                    summary_sheet[key] = self.var_df[
                        'Priority'
                    ].value_counts()[val]
                else:
                    summary_sheet[key] = 0

        # if df is empty add 0 counts for each variant type to summary sheet
        else:
            for cell in count_dict.keys():
                summary_sheet[cell] = 0

        self.var_df.to_excel(
            self.writer, sheet_name="Variants", index=False
        )

        # Set column widths
        ExcelStyles.resize_variant_columns(self, self.workbook["Variants"])

    def create_additional_analysis_page(self):
        '''
        Get Tier3/Null SNVs for Exomiser/deNovo analysis
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        '''
        variant_list = []
        ranked = []
        # Look through Exomiser SNVs and return those that are ranked
        # 1, 2, or 3 and have a score >= 0.75
        for snv in self.wgs_data[self.genome_format][
                self.ex_index
            ][self.genome_data_format]["variants"]:
            ev_to_look_at = []
            for event in snv["reportEvents"]:
                # Filter out mitochondrial + untiered as these are likely
                # artifacts
                if (snv['variantCoordinates']['chromosome'] == 'MT' and
                    event['tier'] is None
                    ):
                    continue
                else:
                    ev_to_look_at.append(event)

            # if we have a list of non MT/untiered events, get highest
            # ranked event from this list + set it as the only report event
            # for that SNV (we do not need the other events now)
            if ev_to_look_at:
                top_event = min(ev_to_look_at, key=lambda x:
                    x['vendorSpecificScores']['rank']
                )
                snv['reportEvents'] = top_event
                ranked.append(snv)

        # We only want Exomiser variants with a score >= 0.75, so we need to
        # filter the list to keep only these
        ranked_and_above_threshold = [
            x for x in ranked if x['reportEvents']['score'] >= 0.75
        ]

        for snv in ranked_and_above_threshold:
            # put reportevents dict within a list to allow it to have an index
            snv['reportEvents'] = [snv['reportEvents']]
            # event index will always be 0 as we have made it so there is only
            # the top ranked event
            event_index = 0
            var_dict = VariantUtils.get_snv_info(
                snv,
                self.proband,
                event_index,
                self.column_list,
                self.mother,
                self.father,
                self.proband_sex
            )
            rank = int(snv['reportEvents'][0]['vendorSpecificScores']['rank'])
            var_dict["Priority"] = f"Exomiser Rank {rank}"
            var_dict["HGVSc"], var_dict["HGVSp"] = (
                VariantNomenclature.get_hgvs_exomiser(
                    snv,
                    self.mane,
                    self.refseq_tsv)
                )
            variant_list.append(var_dict)

        # Get variants with high de novo quality score (these are either SNVs
        # or indels). These variants only appear in the JSON if the quality
        # score is above the threshold
        for snv in self.wgs_data[self.genome_format][
                self.gel_index
            ][self.genome_data_format]["variants"]:
            for event in snv["reportEvents"]:
                if event['segregationPattern'] == 'deNovo':
                    event_index = snv["reportEvents"].index(event)
                    var_dict = VariantUtils.get_snv_info(
                        snv,
                        self.proband,
                        event_index,
                        self.column_list,
                        self.mother,
                        self.father,
                        self.proband_sex
                    )
                    var_dict["Priority"] = "De novo"
                    var_dict["Inheritance"] = "De novo"
                    var_dict["HGVSc"], var_dict["HGVSp"] = (
                        VariantNomenclature.get_hgvs_gel(
                            snv,
                            self.mane,
                            self.refseq_tsv)
                    )
                    variant_list.append(var_dict)

        ex_df = pd.DataFrame(variant_list)
        ex_df = ex_df.drop_duplicates()
        if not ex_df.empty and not self.var_df.empty:
            # Convert all df columns to object type to allow merging without
            # conflicts
            ex_df = ex_df.astype(object)
            self.var_df = self.var_df.astype(object)
            merge_df = ex_df.merge(
                self.var_df,
                on=["Chr", 'Pos', 'Ref', 'Alt'],
                how='left',
                indicator=True,
                suffixes=[None, "_y"]
            )
            # Keep left only == keep only those that are in exomiser df and
            # not in tiered df
            merge_df = merge_df[merge_df['_merge'] == 'left_only']
            # Clean up df by dropping merge column and columns ending _y
            merge_df = merge_df.drop(columns=['_merge'])
            ex_df = merge_df[merge_df.columns.drop(
                list(merge_df.filter(regex='.*\_y'))
            )]

        if not ex_df.empty:
            # Now we have filtered out all variants that are in the GEL tiering
            # page we need to get the top 3 ranks in the exomiser df
            ex_df = VariantUtils.get_top_3_ranked(ex_df)

            # Sort df by priority first, and then gene name
            ex_df = ex_df.sort_values(['Priority', 'Gene'])

        ex_df.to_excel(
            self.writer,
            sheet_name="Extended_analysis",
            index=False
        )

        ExcelStyles.resize_variant_columns(
            self, self.workbook["Extended_analysis"]
        )

        # Add exomiser/de novo variant counts to summary sheet
        summary_sheet = self.workbook["Summary"]
        if 'Priority' in ex_df.columns:
            summary_sheet['B28'] = ex_df['Priority'].str.startswith(
                "De novo"
            ).sum()
            summary_sheet['B27'] = ex_df['Priority'].str.startswith(
                'Exomiser'
            ).sum()

    def write_cnv_reporting_template(self, cnv_sheet_num):
        '''
        Write CNV reporting template to sheet(s) in the workbook.
        Inputs:
            cnv_sheet_num (int): number to append to CNV title
        Outputs:
            None, adds content to openpxyl workbook
        '''
        cnv = self.workbook.create_sheet(f"cnv_interpret_{cnv_sheet_num}")
        titles = {
            "Intragenic CNVs should be analysed using SNV guidelines": [1,2], 
            "Chromosomal region/gene": [3, 2],
            "Start": [3, 3],
            "Stop": [3, 4],
            "Gain/Loss": [3, 5],
            "Zygosity": [3, 6],
            "Gene Content:": [7, 2],
            "Evidence": [6, 3],
            "Possible evidence": [6, 7],
            "Prevalence in controls": [11, 2],
            "Microdel/dup syndromes": [13, 2],
            "Literature search": [15, 2],
            "Inheritance in this case": [18, 2],
            "ACMG/ACGS Classification": [20, 2],
        }

        content = {
            "Does the CNV contain protein coding genes? How many?": [8,2], 
            "OMIM/green genes?": [9,2], 
            "Any disease genes relevant to phenotype?": [10,2], 
            "Are similar CNVs in the gnomAD-SV database? Or in DGV?": [12,2], 
            "Does this CNV overlap with a known microdeletion or "
            "microduplication syndrome? Check decipher, pubmed, new " 
            "ACMG CNV guidelines Table S3": [14,2], 
            "Similar CNVs in HGMD, decipher, pubmed listed as pathogenic?"
            "Are they de novo? Do they segregate with disease in the reported"
            "family?": [16, 2], 
            "Does gene of interest have evidence of HI/TS?": [17,2], 
            "In this case is the CNV de novo, inherited, unknown? Good "
            "phenotype fit? Non-segregation in affected family"
            "members?": [19, 2],
            "1A/3": [7, 7],
            "2G/2F for loss,\n2C-2G for gain": [11, 7],
            "2A-2B": [13, 7],
            "4A-4K": [15, 7],
            "2A, 2B, 2H-2K for loss, \n2A-2E for gain": [17, 7],
            "5A-5H": [18, 7],
            "CNV pathogenicity calculators:": [2, 9],
            "Loss:": [3, 9],
            "Gain:": [4, 9]
        }
        for key, val in titles.items():
            cnv.cell(val[0], val[1]).value = key
            cnv.cell(val[0], val[1]).font = Font(
                bold=True, name=DEFAULT_FONT.name
            )

        for key, val in content.items():
            cnv.cell(val[0], val[1]).value = key

        # Add pathogenicity calculator links
        cnv['J3'].hyperlink = "https://cnvcalc.clinicalgenome.org/cnvcalc/"\
            "cnv-loss"
        cnv['J4'].hyperlink = "https://cnvcalc.clinicalgenome.org/cnvcalc/"\
            "cnv-gain"

        cnv.column_dimensions['B'].width = 35
        cnv.column_dimensions['G'].width = 20
        for col in ['C', 'D', 'E', 'F']:
            cnv.column_dimensions[col].width = 15

        for i in [8, 9, 10, 12, 14, 16, 17, 19]:
            cnv.row_dimensions[i].height = 60
            cnv[f"B{i}"].alignment = Alignment(
                        wrapText=True, vertical="center"
                )
        cnv.row_dimensions[9].height = 15

        # merge evidence cells
        merge_dict = {7: 10, 11: 12, 13: 14, 15: 16, 18: 19}
        for start, end in merge_dict.items():
            cnv.merge_cells(range_string=f'G{start}:G{end}')

        for row in range(6, 20):
            cnv.merge_cells(
                start_row=row, end_row=row, start_column=3, end_column=6)

        # define which rows should have borders
        row_ranges = {
            'horizontal': [
                'B3:D3', 'B4:F4', 'B6:G6', 'B7:G7', 'B8:G8', 'B9:G9',
                'B10:G10', 'B11:G11',
                'B12:G12', 'B13:G13', 'B14:G14', 'B15:G15', 'B16:G16',
                'B17:G17', 'B18:G18', 'B19:G19', 'B20:G20',
            ],
            'horizontal_thick': [
                'B3:F3', 'B5:F5', 'B6:G6', 'B7:G7', 'B20:G20', 'B21:G21'
            ],
            'vertical': [
                'E2:E3', 'G6:G20'
            ],
            'vertical_thick': [
                'B3:B4', 'B6:B20', 'G3:G4', 'C6:C20', 'H6:H20'
            ]
        }

        ExcelStyles.borders(self, row_ranges, cnv)

        # add some colour
        colour_cells = {
            'FFC000': ['G17'],
            'FFFF00': ['G15', 'G18'],
            '0070C0': ['G13'],
            '00B050': ['G11'],
            'D9D9D9': ['G7']
        }
        ExcelStyles.colours(self, colour_cells, cnv)

        # align text
        for row in [7, 11, 13, 15, 17, 18]:
            cnv[f"G{row}"].alignment = Alignment(
                wrapText=True, vertical="center", horizontal="center"
            )

    def write_snv_reporting_template(self, report_sheet_num) -> None:
        """
        Writes sheet(s) to Excel file with formatting for reporting against
        ACMG criteria
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        """
        report = self.workbook.create_sheet(
            f"snv_interpret_{report_sheet_num}"
        )

        titles = {
            "Gene": [2, 2],
            "HGVSc": [2, 3],
            "HGVSp": [2, 4],
            "EVIDENCE": [8, 3],
            "PATHOGENIC": [8, 7],
            "P_STRENGTH": [8, 8],
            "P_POINTS": [8, 9],
            "BENIGN": [8, 10],
            "B_STRENGTH": [8, 11],
            "B_POINTS": [8, 12],
            "Associated disease": [4, 2],
            "Known inheritance": [5, 2],
            "Prevalence": [6, 2],
            ("Allele frequency is >5% (or gene-specific cut off) in "
             "population data e.g. gnomAD, UKB"): [9, 2],
            ("Null variant in a gene where LOF is known mechanism "
             "of disease\nand non-canonical splice variants where "
             "RNA analysis confirms\naberrant transcription"): [10, 2],
            ("Same AA change as previously established pathogenic "
             "variant\nregardless of nucleotide change and splicing "
             "variants within\nsame motif with identical predicted "
             "effect"): [11, 2],
            ("De novo (confirmed) / observed in\nhealthy adult "
             "with full penetrance expected at an early age"): [12, 2],
            "In vivo / in vitro functional studies": [13, 2],
            "Prevalence in affected > controls": [14, 2],
            ("In mutational hot spot and/or critical functional "
             "domain, without\nbenign variation"): [15, 2],
            ("Freq in controls eg gnomAD, low/absent (PM2) or allele "
             "frequency is greater than expected for disorder (BS1)"): [16, 2],
            "Detected in trans/in cis with pathogenic variant": [17, 2],
            ("In frame protein length change/stop-loss variants, "
             "non repeat\nvs. repeat region"): [18, 2],
            ("Missense change at AA where different likely/pathogenic\n"
             "missense change seen before"): [19, 2],
            "Assumed de novo (no confirmation)": [20, 2],
            "Cosegregation with disease in family, not in unaffected": [21, 2],
            ("Missense where low rate of benign missense and common\n"
             "mechanism (Z score ≥3.09), or missense where LOF common\n"
             "mechanism"): [22, 2],
            "Multiple lines of computational evidence": [23, 2],
            ("Phenotype/FH specific for disease of single etiology, or\n"
             "alternative genetic cause of disease detected"): [24, 2],
            ("Synonymous change, no affect on splicing, not conserved; "
             "splice\nvariants confirmed to have no impact"): [25, 2],
            "POINTS": [26, 7],
        }
        for key, val in titles.items():
            report.cell(val[0], val[1]).value = key
            report.cell(val[0], val[1]).font = Font(
                bold=True, name=DEFAULT_FONT.name
            )
        classifications = {
            "PVS1": [(10, 7)],
            "PS1": [(11, 7)],
            "PS2": [(12, 7)],
            "PS3": [(13, 7)],
            "PS4": [(14, 7)],
            "PM1": [(15, 7)],
            "PM2": [(16, 7)],
            "PM3": [(17, 7)],
            "PM4": [(18, 7)],
            "PM5": [(19, 7)],
            "PM6": [(20, 7)],
            "PP1": [(21, 7)],
            "PP2": [(22, 7)],
            "PP3": [(23, 7)],
            "PP4": [(24, 7)],
            "BA1": [(9, 10)],
            "BS2": [(12, 10)],
            "BS3": [(13, 10)],
            "BS1": [(16, 10)],
            "BP2": [(17, 10)],
            "BP3": [(18, 10)],
            "BS4": [(21, 10)],
            "BP1": [(22, 10)],
            "BP4": [(23, 10)],
            "BP5": [(24, 10)],
            "BP7": [(25, 10)]
        }

        for key, values in classifications.items():
            for val in values:
                report.cell(val[0], val[1]).value = key

        # nice formatting of title text and columns
        for col in (['B', 'C', 'G', 'H', 'I', 'J', 'K', 'L']):
            for row in range(8, 26):
                if row == 8:
                    report[f"{col}{row}"].alignment = Alignment(
                           wrapText=True, vertical="center",
                           horizontal="center"
                    )
                elif (col == "B" and row != 8) or (col == "C" and row != 8):
                    report.row_dimensions[row].height = 50
                    report[f"{col}{row}"].alignment = Alignment(
                           wrapText=True, vertical="center"
                    )
                else:
                    report[f"{col}{row}"].alignment = Alignment(
                           wrapText=True, vertical="center",
                           horizontal="center"
                    )
                    report[f"{col}{row}"].font = Font(size=14,
                                                      name=DEFAULT_FONT.name)

        for col in (['B', 'C', 'D']):
            for row in range(2, 7):
                report.row_dimensions[row].height = 20
                report[f"{col}{row}"].font = Font(size=14, bold=True,
                                                  name=DEFAULT_FONT.name)

        # merge associated disease, inheritance and prevalence cells
        for row in range(4, 8):
            report.merge_cells(
                start_row=row, end_row=row, start_column=3, end_column=12)

        # merge evidence cells
        for row in range(8, 27):
            report.merge_cells(
                start_row=row, end_row=row, start_column=3, end_column=6)

        # merge POINTS cells
        report.merge_cells(
                start_row=26, end_row=26, start_column=8, end_column=12)

        # set appropriate widths
        report.column_dimensions['B'].width = 62
        report.column_dimensions['C'].width = 35
        report.column_dimensions['D'].width = 35
        report.column_dimensions['E'].width = 5
        report.column_dimensions['F'].width = 5
        report.column_dimensions['G'].width = 14
        report.column_dimensions['H'].width = 14
        report.column_dimensions['I'].width = 14
        report.column_dimensions['J'].width = 14
        report.column_dimensions['K'].width = 14
        report.column_dimensions['L'].width = 14

        # do some colouring
        colour_cells = {
            'E46C0A': ['G11', 'G12', 'G13', 'G14'],
            'FFC000': ['G15', 'G16', 'G17', 'G18', 'G19', 'G20'],
            'FFFF00': ['G21', 'G22', 'G23', 'G24'],
            '00B0F0': ['J12', 'J13', 'J16', 'J21'],
            '92D050': ['J17', 'J18', 'J22', 'J23', 'J24', 'J25'],
            '0070C0': ['J9'],
            'FF0000': ['G10'],
            'D9D9D9': ['G9', 'G25', 'H9', 'H25', 'I9', 'I25',
                       'J10', 'J11', 'J14', 'J15', 'J19', 'J20',
                       'K10', 'K11', 'K14', 'K15', 'K19', 'K20',
                       'L10', 'L11', 'L14', 'L15', 'L19', 'L20'
                       ]

        }

        ExcelStyles.colours(self, colour_cells, report)

        # add some borders
        row_ranges = {
            'horizontal': [
                'B3:D3', 'B4:J4', 'B5:L5',
                'B6:L6', 'B7:L7', 'B8:L8', 'B9:L9', 'B10:L10', 'B11:L11',
                'B12:L12', 'B13:L13', 'B14:L14', 'B15:L15', 'B16:L16',
                'B17:L17', 'B18:L18', 'B19:L19', 'B20:L20', 'B21:L21',
                'B22:L22', 'B23:L23', 'B24:L24', 'B25:L25'
            ],
            'horizontal_thick': [
                'B2:D2', 'B4:L4', 'B7:L7', 'B8:L8', 'B26:L26', 'B27:L27'
            ],
            'vertical': [
                'E2:E3', 'G8:G26', 'H8:H25', 'I8:I25', 'J8:J25',
                'K8:K25', 'L8:L25',
            ],
            'vertical_thick': [
                'B2:B6', 'B8:B26', 'C2:C6', 'C8:C26', 'M4:M6', 'M8:M26',
                'E2:E3'
            ]
        }
        ExcelStyles.borders(self, row_ranges, report)
