#!/usr/bin/env python3
from openpyxl.styles import Alignment, Border, DEFAULT_FONT, Font, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook

# openpyxl style settings
THIN = Side(border_style="thin", color="000000")
MEDIUM = Side(border_style="medium", color="000001")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


class ExcelStyles():
    '''
    Functions to add style to excel workbook
    '''
    def colours(self, colour_cells, sheet):
        '''
        Add colour to cells in workbook.
        Inputs:
            colour_cells (dict): dict of colour number keys with cells that
            should be that colour as values
            sheet (str): sheet in workbook to add colour to
        Outputs:
            None, adds content to openpxyl workbook
        '''
        for colour, cells in colour_cells.items():
            for cell in cells:
                sheet[cell].fill = PatternFill(
                    patternType="solid", start_color=colour
                )

    def borders(self, row_ranges, sheet):
        '''
        Add borders to sheet.
        Inputs:
            row_ranges (dict): dict of border type and rows/columns that should
            have that border
            sheet (str): sheet in workbook to add borders to
        Outputs:
            None, adds content to openpxyl workbook
        '''
        for side, values in row_ranges.items():
            for row in values:
                for cells in sheet[row]:
                    for cell in cells:
                        # border style is immutable => copy current and modify
                        cell_border = cell.border.copy()
                        if side == 'horizontal':
                            cell_border.top = THIN
                        if side == 'horizontal_thick':
                            cell_border.top = MEDIUM
                        if side == 'vertical':
                            cell_border.left = THIN
                        if side == 'vertical_thick':
                            cell_border.left = MEDIUM
                        cell.border = cell_border

    def resize_variant_columns(self, sheet):
        '''
        Resize columns for both gel tiering variant page + exomiser page to the
        same width
        Inputs:
            sheet: openpxyl sheet on which to resize the columns
        Outputs:
            None, adds content to openpxyl workbook
        '''
        for col in ['C', 'D']:
            sheet.column_dimensions[col].width = 14
        for col in ['J']:
            sheet.column_dimensions[col].width = 10
        for col in ['F', 'G', 'H', 'I', 'L']:
            sheet.column_dimensions[col].width = 6
        for col in ['R', 'S', 'T']:
            sheet.column_dimensions[col].width = 12
        sheet.column_dimensions['B'].width = 5
        sheet.column_dimensions['O'].width = 15
        for col in ['M', 'N', 'P', 'Q', 'U']:
            sheet.column_dimensions[col].width = 20
        for col in ['V', 'W']:
            sheet.column_dimensions[col].width = 25


class DropDown():
    '''
    Handle adding drop down menus to Excel workbook.
    '''
    def drop_down(self) -> None:
        """
        Function to add drop-downs in the report tab for entering
        ACMG criteria for classification, as well as a boolean
        drop down into the additional 'Interpreted' column of
        the variant sheet(s).
        Inputs:
            None
        Outputs:
            None, adds content to openpxyl workbook
        """
        wb = load_workbook(filename=self.args.output_filename)

        # adding dropdowns in report table
        for sheet_num in range(1, self.args.acmg+1):
            # adding strength dropdown except for BA1
            report_sheet = wb[f"snv_interpret_{sheet_num}"]
            cells_for_strength = ['H10', 'H11', 'H12', 'H13', 'H14', 'H15',
                                  'H16', 'H17', 'H18', 'H19', 'H20', 'H21',
                                  'H22', 'H23', 'H24', 'K12', 'K13', 'K16',
                                  'K17', 'K18', 'K21', 'K22', 'K23', 'K24',
                                  'K25']
            strength_options = '"Very Strong, Strong, Moderate, \
                                 Supporting, NA"'
            DropDown.get_drop_down(
                self,
                dropdown_options=strength_options,
                prompt='Select from the list',
                title='Strength',
                sheet=report_sheet,
                cells=cells_for_strength
            )

            # add stregth for BA1
            BA1_options = '"Stand-Alone, Very Strong, Strong, Moderate, \
                            Supporting, NA"'
            DropDown.get_drop_down(
                self, dropdown_options=BA1_options,
                prompt='Select from the list',
                title='Strength',
                sheet=report_sheet,
                cells=['K9']
            )

            # adding final classification dropdown
            report_sheet['B26'] = 'FINAL ACMG CLASSIFICATION'
            report_sheet['B26'].font = Font(bold=True, name=DEFAULT_FONT.name)
            class_options = '"Pathogenic,Likely Pathogenic, \
                              Uncertain Significance, \
                              Likely Benign, Benign"'
            DropDown.get_drop_down(
                self,
                dropdown_options=class_options,
                prompt='Select from the list',
                title='ACMG classification',
                sheet=report_sheet,
                cells=['C26']
            )

        wb.save(self.args.output_filename)

    def get_drop_down(self, dropdown_options, prompt, title, sheet, cells):
        """
        create the drop-downs items for designated cells
        Inputs:
            dropdown_options: str containing drop-down items
            prompt: prompt message for drop-down
            title: title message for drop-down
            sheet: openpyxl.Writer writer object current worksheet
            cells: list of cells to add drop-down
        Outputs:
            None, adds content to openpxyl workbook
        """
        options = dropdown_options
        val = DataValidation(type='list', formula1=options,
                             allow_blank=True)
        val.prompt = prompt
        val.promptTitle = title
        sheet.add_data_validation(val)
        for cell in cells:
            val.add(sheet[cell])
        val.showInputMessage = True
        val.showErrorMessage = True
